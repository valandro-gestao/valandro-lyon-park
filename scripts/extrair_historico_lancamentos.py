"""
Extrai o histórico mensal (faturamento, resultado, repasse) da planilha
histórica original ("Lyon - Dados para Relatórios.xlsx") e grava em
migrations/data/historico_lancamentos.json — o único formato que a migração
0002 lê. Executado uma única vez, localmente; não faz parte do runtime da
aplicação e não é uma funcionalidade do usuário (sprint v1.1.2, requisito 1).

Cada aba tem um layout próprio (mês em linha, mês em coluna, ou os dois
contratantes do Pátio lado a lado). Todo índice de coluna abaixo foi
verificado programaticamente contra o cabeçalho real da planilha (nunca
contado à mão) — e o próprio script reverifica o rótulo esperado em cada
coluna antes de extrair, reportando como "problema" (não como dado) se o
rótulo não bater. Essa verificação existe porque scripts/importar_historico.py
tinha pelo menos um mapeamento de coluna incorreto (W-Tower Caxias, corrigido
separadamente na migração 0003) — aqui a suposição nunca fica implícita.

Regra de reconstrução aplicada uniformemente (documentada, não silenciosa):
  aluguel_calculado = max(0, valor da planilha)
Motivo: em unidades com prejuízo acumulado (COM_ALIQUOTA_CUMUL), a planilha
histórica registra o valor bruto da fórmula (percentual × resultado com
saldo negativo), que fica negativo por vários meses até o prejuízo ser
compensado — exatamente o que app/calculators/cumulativo.py também calcula
internamente antes de aplicar `if resultado_com_prejuizo > 0 else aluguel=0`.
Sem esse piso, o comparativo do PDF mostraria "Repasse: -R$ X" em meses de
prejuízo, o que não corresponde a nenhum valor real cobrado.

`resultado` é gravado como está na planilha (pode ser negativo — é assim que
o próprio ResultadoUnidade.resultado já se comporta hoje para essas mesmas
unidades).

Park Tower e Monza não têm coluna própria de "Resultado" — reconstrução por
fórmula conhecida (a mesma que o calculador usa), não invenção de dado:
  Park Tower (COM_ALIQUOTA, PE=0): resultado == subtotal
  Monza      (COM_FAIXAS, sem PE/alíquota): resultado == faturamento

Uso:
  .venv/bin/python scripts/extrair_historico_lancamentos.py
"""
import json
import os
import sys
import datetime as dt

import openpyxl

EXCEL = os.path.expanduser("~/Downloads/Lyon - Dados para Relatórios.xlsx")
SAIDA = os.path.join(os.path.dirname(__file__), "..", "migrations", "data", "historico_lancamentos.json")

# Corte: a partir daqui o histórico já é alimentado pela própria ferramenta
# (lancamentos já tem linhas reais desde 2026-05). Extraído mesmo assim —
# a migração decide o que já existe — mas não vamos além do que a planilha
# efetivamente tem.
MES_CORTE = "2026-05"

COL_MES = 1  # coluna B em todas as abas "mês em linha"


def _mes_ref(data) -> str | None:
    if not isinstance(data, (dt.datetime, dt.date)):
        return None
    return f"{data.year:04d}-{data.month:02d}"


def _num(v) -> float | None:
    if v is None:
        return None
    try:
        f = float(v)
        return f if f == f else None  # descarta NaN
    except (TypeError, ValueError):
        return None


def _floor0(v: float | None) -> float:
    v = v or 0.0
    return v if v > 0 else 0.0


def _checar_header(ws, linha_header: int, indice: int, esperado: str) -> str | None:
    """Confirma que a coluna/linha realmente contém o rótulo esperado antes
    de extrair. Retorna None se ok, ou uma mensagem de erro."""
    header = list(next(ws.iter_rows(min_row=linha_header, max_row=linha_header, values_only=True)))
    real = header[indice] if indice < len(header) else None
    if not isinstance(real, str) or esperado.lower() not in real.lower():
        return f"esperava rótulo contendo {esperado!r} na posição {indice}, encontrado {real!r}"
    return None


# ─── abas "mês em linha" ──────────────────────────────────────────────────
# uid -> (aba, col_fat, rotulo_fat, col_res, rotulo_res, col_alug1, rotulo_alug1, col_alug2_ou_None, rotulo_alug2)
MAPA_LINHA = {
    "fiergs":          ("Fiergs",          2, "Faturamento", 10, "Resultado", 13, "Aluguel", None, None),
    "anitta_mall":     ("Anitta Mall",     2, "Faturamento", 10, "Resultado", 12, "Aluguel", 13, "Aluguel"),
    "a_schneider":     ("A. Schneider",    2, "Faturamento", 6,  "Resultado", 8,  "Aluguel", None, None),
    "axis":            ("Axis",            2, "Faturamento", 4,  "Resultado", 6,  "Aluguel", None, None),
    "dom_pedro":       ("Dom Pedro",       2, "Faturamento", 6,  "Resultado", 8,  "Aluguel", None, None),
    "fk":              ("FK",              2, "Faturamento", 6,  "Resultado", 8,  "Aluguel", None, None),
    "ilp":             ("ILP",             2, "Faturamento", 10, "Resultado", 12, "Aluguel", None, None),
    "in_1183":         ("In 1183",         2, "Faturamento", 9,  "Resultado", 11, "Aluguel", None, None),
    "mw_tristeza":     ("MW Tristeza",     2, "Faturamento", 9,  "Resultado", 12, "Aluguel", None, None),
    "nl_2800":         ("NL 2800",         2, "Faturamento", 4,  "Resultado", 8,  "Aluguel", None, None),
    "praia_de_bellas": ("Praia de Bellas", 2, "Faturamento", 6,  "Resultado", 10, "Repasse", None, None),
    "vasco":           ("Vasco",           2, "Faturamento", 4,  "Resultado", 6,  "Aluguel", None, None),
    "viva_trindade":   ("Viva Trindade",   2, "Faturamento", 9,  "Resultado", 12, "Aluguel", None, None),
    "w_tower_caxias":  ("W-Tower Caxias",  2, "Faturamento", 10, "Resultado", 12, "Aluguel", None, None),
    "ekos":            ("EKOS",            2, "Faturamento", 11, "Resultado", 13, "Aluguel", None, None),
    "oka":             ("OKA",             2, "Faturamento", 11, "Resultado", 13, "Aluguel", None, None),
}

# resultado derivado por fórmula (não há coluna própria na planilha)
MAPA_LINHA_RESULTADO_DERIVADO = {
    "park_tower": ("Park Tower", 2, "Faturamento", 4, "Subtotal", 6, "Aluguel"),
    "monza":      ("Monza",      2, "Faturamento", None, None,    4, "Aluguel"),
}


def _extrair_linha(wb, aba, col_fat, rot_fat, col_res, rot_res, col_al1, rot_al1, col_al2, rot_al2):
    if aba not in wb.sheetnames:
        return [], [f"aba '{aba}' não encontrada"]
    ws = wb[aba]
    erros = []
    for col, rot in ((col_fat, rot_fat), (col_res, rot_res), (col_al1, rot_al1), (col_al2, rot_al2)):
        if col is None:
            continue
        e = _checar_header(ws, 3, col, rot)
        if e:
            erros.append(e)
    if erros:
        return [], erros

    registros = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        mes = _mes_ref(row[COL_MES] if len(row) > COL_MES else None)
        if mes is None or mes >= MES_CORTE:
            continue
        fat = _num(row[col_fat]) if col_fat < len(row) else None
        if fat is None or fat == 0:
            continue
        resultado = _num(row[col_res]) if col_res is not None and col_res < len(row) else None
        aluguel = _floor0(_num(row[col_al1])) if col_al1 < len(row) else 0.0
        if col_al2 is not None and col_al2 < len(row):
            aluguel += _floor0(_num(row[col_al2]))
        registros.append({
            "mes_referencia": mes,
            "faturamento": round(fat, 2),
            "resultado": round(resultado, 2) if resultado is not None else None,
            "aluguel_calculado": round(aluguel, 2),
        })
    return registros, []


def _extrair_linha_resultado_derivado(wb, aba, col_fat, rot_fat, col_res_fonte, rot_res_fonte, col_alug, rot_alug):
    if aba not in wb.sheetnames:
        return [], [f"aba '{aba}' não encontrada"]
    ws = wb[aba]
    erros = []
    checagens = [(col_fat, rot_fat), (col_alug, rot_alug)]
    if col_res_fonte is not None:
        checagens.append((col_res_fonte, rot_res_fonte))
    for col, rot in checagens:
        e = _checar_header(ws, 3, col, rot)
        if e:
            erros.append(e)
    if erros:
        return [], erros

    registros = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        mes = _mes_ref(row[COL_MES] if len(row) > COL_MES else None)
        if mes is None or mes >= MES_CORTE:
            continue
        fat = _num(row[col_fat]) if col_fat < len(row) else None
        if fat is None or fat == 0:
            continue
        resultado = fat if col_res_fonte is None else (
            _num(row[col_res_fonte]) if col_res_fonte < len(row) else None
        )
        aluguel = _floor0(_num(row[col_alug])) if col_alug < len(row) else 0.0
        registros.append({
            "mes_referencia": mes,
            "faturamento": round(fat, 2),
            "resultado": round(resultado, 2) if resultado is not None else None,
            "aluguel_calculado": round(aluguel, 2),
        })
    return registros, []


# ─── abas "transpostas" (indicador em linha, mês em coluna) ──────────────
# Medcenter e Viva Open Mall: linha 2 tem os nomes dos meses; colunas cujo
# cabeçalho é um número (2018-2030) são totais anuais e servem apenas para
# indicar a virada de ano — não representam um mês.
# uid -> (aba, row_fat, rotulo_fat, row_resultado, rotulo_resultado, row_repasse, rotulo_repasse)
MAPA_TRANSPOSTO = {
    "medcenter":      ("Medcenter",      3, "Faturamento", 14, "Resultado", 18, "Saldo a pagar"),
    "viva_open_mall": ("Viva Open Mall", 3, "Faturamento", 17, "Resultado", 23, "Saldo a pagar"),
}

MESES_PT = ["janeiro", "fevereiro", "março", "abril", "maio", "junho",
            "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]


def _extrair_transposto(wb, aba, row_fat, rot_fat, row_res, rot_res, row_alug, rot_alug):
    if aba not in wb.sheetnames:
        return [], [f"aba '{aba}' não encontrada"]
    ws = wb[aba]
    max_linha = max(row_fat, row_res, row_alug)
    linhas = list(ws.iter_rows(min_row=1, max_row=max_linha, values_only=True))

    erros = []
    for row_idx, rot in ((row_fat, rot_fat), (row_res, rot_res), (row_alug, rot_alug)):
        label = linhas[row_idx - 1][1] if len(linhas[row_idx - 1]) > 1 else None
        if not isinstance(label, str) or rot.lower() not in label.lower():
            erros.append(f"linha {row_idx}: esperava rótulo contendo {rot!r}, encontrado {label!r}")
    if erros:
        return [], erros

    header = linhas[1]  # linha 2: nomes de mês, ou ano (int/float) marcando a virada
    fat_row = linhas[row_fat - 1]
    res_row = linhas[row_res - 1]
    alug_row = linhas[row_alug - 1]

    registros = []
    for col_idx in range(2, len(header)):
        h = header[col_idx]
        if not isinstance(h, str) or h.strip().lower() not in MESES_PT:
            continue
        mes_num = MESES_PT.index(h.strip().lower()) + 1

        ano_col = None
        for j in range(col_idx, len(header)):
            if isinstance(header[j], (int, float)):
                ano_col = int(header[j])
                break
        if ano_col is None:
            continue  # sem marcador de ano à direita — não há como inferir com segurança

        fat = _num(fat_row[col_idx]) if col_idx < len(fat_row) else None
        if fat is None or fat == 0:
            continue
        mes_referencia = f"{ano_col:04d}-{mes_num:02d}"
        if mes_referencia >= MES_CORTE:
            continue
        resultado = _num(res_row[col_idx]) if col_idx < len(res_row) else None
        aluguel = _floor0(_num(alug_row[col_idx])) if col_idx < len(alug_row) else 0.0
        registros.append({
            "mes_referencia": mes_referencia,
            "faturamento": round(fat, 2),
            "resultado": round(resultado, 2) if resultado is not None else None,
            "aluguel_calculado": round(aluguel, 2),
        })

    dedup = {r["mes_referencia"]: r for r in registros}  # a última leitura de um mês prevalece
    return sorted(dedup.values(), key=lambda r: r["mes_referencia"]), []


# ─── Pátio (REAL / MAIOJAMA lado a lado na mesma aba) ────────────────────
_PATIO_COLS = {
    "fat_real": (7, "Faturamento REAL"), "res_real": (12, "Resultado REAL"),
    "alug_real": (14, "Aluguel"), "outros_real": (15, "Outros Serviços REAL"),
    "fat_maio": (18, "Faturamento MAIOJAMA"), "res_maio": (23, "Resultado MAIOJAMA"),
    "alug_maio": (25, "Aluguel"), "outros_maio": (26, "Outros Serviços MAIOJAMA"),
}


def _extrair_patio(wb):
    aba = "Patio"
    if aba not in wb.sheetnames:
        msg = f"aba '{aba}' não encontrada"
        return {}, {"patio_real": [msg], "patio_maiojama": [msg]}
    ws = wb[aba]

    erros = []
    for nome, (col, rot) in _PATIO_COLS.items():
        e = _checar_header(ws, 3, col, rot)
        if e:
            erros.append(f"{nome}: {e}")
    if erros:
        return {}, {"patio_real": erros, "patio_maiojama": erros}

    c = {k: v[0] for k, v in _PATIO_COLS.items()}
    saida = {"patio_real": [], "patio_maiojama": []}
    for row in ws.iter_rows(min_row=4, values_only=True):
        mes = _mes_ref(row[COL_MES] if len(row) > COL_MES else None)
        if mes is None or mes >= MES_CORTE:
            continue

        fat_real = _num(row[c["fat_real"]]) if len(row) > c["fat_real"] else None
        if fat_real:
            res_real = _num(row[c["res_real"]]) if len(row) > c["res_real"] else None
            alug_real = _floor0(_num(row[c["alug_real"]])) if len(row) > c["alug_real"] else 0.0
            outros_real = _floor0(_num(row[c["outros_real"]])) if len(row) > c["outros_real"] else 0.0
            saida["patio_real"].append({
                "mes_referencia": mes,
                "faturamento": round(fat_real, 2),
                "resultado": round(res_real, 2) if res_real is not None else None,
                "aluguel_calculado": round(alug_real, 2),
                "extras": {"repasse_outros": round(outros_real, 2)} if outros_real else {},
            })

        fat_maio = _num(row[c["fat_maio"]]) if len(row) > c["fat_maio"] else None
        if fat_maio:
            res_maio = _num(row[c["res_maio"]]) if len(row) > c["res_maio"] else None
            alug_maio = _floor0(_num(row[c["alug_maio"]])) if len(row) > c["alug_maio"] else 0.0
            outros_maio = _floor0(_num(row[c["outros_maio"]])) if len(row) > c["outros_maio"] else 0.0
            saida["patio_maiojama"].append({
                "mes_referencia": mes,
                "faturamento": round(fat_maio, 2),
                "resultado": round(res_maio, 2) if res_maio is not None else None,
                "aluguel_calculado": round(alug_maio, 2),
                "extras": {"repasse_outros": round(outros_maio, 2)} if outros_maio else {},
            })
    return saida, {}


def main():
    if not os.path.exists(EXCEL):
        sys.exit(f"Planilha não encontrada em {EXCEL}. Este script só roda localmente, uma vez.")

    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)

    resultado_final: dict[str, list[dict]] = {}
    problemas: dict[str, list[str]] = {}

    for uid, (aba, col_fat, rot_fat, col_res, rot_res, col_al1, rot_al1, col_al2, rot_al2) in MAPA_LINHA.items():
        registros, erros = _extrair_linha(wb, aba, col_fat, rot_fat, col_res, rot_res, col_al1, rot_al1, col_al2, rot_al2)
        if erros:
            problemas[uid] = erros
        resultado_final[uid] = registros

    for uid, (aba, col_fat, rot_fat, col_res, rot_res, col_alug, rot_alug) in MAPA_LINHA_RESULTADO_DERIVADO.items():
        registros, erros = _extrair_linha_resultado_derivado(wb, aba, col_fat, rot_fat, col_res, rot_res, col_alug, rot_alug)
        if erros:
            problemas[uid] = erros
        resultado_final[uid] = registros

    for uid, (aba, row_fat, rot_fat, row_res, rot_res, row_alug, rot_alug) in MAPA_TRANSPOSTO.items():
        registros, erros = _extrair_transposto(wb, aba, row_fat, rot_fat, row_res, rot_res, row_alug, rot_alug)
        if erros:
            problemas[uid] = erros
        resultado_final[uid] = registros

    patio_dados, patio_problemas = _extrair_patio(wb)
    resultado_final.update(patio_dados)
    problemas.update(patio_problemas)

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    with open(SAIDA, "w", encoding="utf-8") as f:
        json.dump(resultado_final, f, ensure_ascii=False, indent=2, sort_keys=True)

    print(f"Gravado em {SAIDA}")
    print()
    for uid, registros in sorted(resultado_final.items()):
        if not registros:
            extra = f" ({'; '.join(problemas[uid])})" if uid in problemas else ""
            print(f"  {uid:20s} SEM DADOS{extra}")
            continue
        print(f"  {uid:20s} {registros[0]['mes_referencia']} -> {registros[-1]['mes_referencia']}  "
              f"({len(registros)} competências)")
    if problemas:
        print()
        print("Problemas:")
        for uid, msgs in sorted(problemas.items()):
            for msg in msgs:
                print(f"  {uid}: {msg}")


if __name__ == "__main__":
    main()
